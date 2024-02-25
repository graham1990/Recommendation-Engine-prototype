import csv
import io
import itertools as it
import openpyxl
import pandas as pd
from typing import Any, Union


class DataDesc:
    """The data and meta data describing the data"""

    def __init__(self, **meta):
        self._data = None
        self._meta = meta

    @property
    def data(self) -> Any:
        return self._data

    @property
    def meta(self) -> dict:
        return self._meta

    def meta_update(self, new_attrs: dict) -> None:
        """Add new meta data and replace any attr names already present"""
        self._meta.update(new_attrs)


class ColumnKey:
    """Column data identified by Key"""

    def __init__(self, data: dict = None):
        self._data = data if data else dict()

    def __setitem__(self, __key: Any, __value: Any) -> None:
        self._data[__key] = __value

    def __getitem__(self, __key: Any) -> Any:
        return self._data[__key]
    
    def __contains__(self, __key: Any) -> bool:
        return self._data.__contains__(__key)

    def __str__(self) -> str:
        return str(self._data)

    def __repr__(self) -> str:
        return repr(self._data)


class StreamUtil:
    def to_column_key(self, remove_missing: bool = True) -> list[ColumnKey]:
        """Convert the loaded spreadsheet to dict where the column name is the key and the data is the value"""
        data = []
        for row_no in range(1, self.row_count + 1):
            params = ColumnKey()
            for column_no in range(1, self.column_count + 1):
                param_name = self.get_column_name(column=column_no)
                value = self.get_row_column_contents(row=row_no, column=column_no)
                if remove_missing:
                    param_name = param_name.strip() if isinstance(param_name, str) else param_name
                    params[param_name] = (
                        self.__class__._preprocess_application_data(
                            value.strip(), self._missing_values
                        )
                        if isinstance(value, str)
                        else value
                    )
                else:
                    params[param_name.strip()] = (
                        value.strip() if isinstance(value, str) else value
                    )

            data.append(params)
        return data


class Local_Excel_Workbook_Stream(StreamUtil):
    """Local loading and writing of Excel 2010 format files Stream.
    Data table is extracted from the default tab or the named tab"""

    def __init__(self):
        self._shape = {}

    def load_data(self, xlsx_name: str, **options) -> None:
        """Load a workbook"""
        self._missing_values: list[str] = options.get("missing_values", [])
        with open(xlsx_name, "rb") as excel_file:
            self._contents = excel_file.read()

    def save_data(self, xlsx_name):
        """Save the contents to an xlsx Excel file"""
        with open(xlsx_name, "wb") as excel_file:
            excel_file.write(self._contents)

    @staticmethod
    def _preprocess_application_data(data: Any, missing_values: list[str]) -> Any:
        """Changes the data to empty if any missing_values are found"""
        if data in missing_values:
            return None
        return data

    def get_column_name(self, column: int) -> str:
        """Get the column name from row 1 (1 based), column_no (1 based)"""
        return self._sheet_obj.cell(row=1, column=column).value

    def get_row_column_contents(self, row: int, column: int) -> Union[str, int]:
        # def get_row_column_contents(self, row: int, column: int) -> str | int:
        """Get the value from row (1 based), column (1 based)"""
        # Start from the data row, not header
        return self._sheet_obj.cell(row=row + 1, column=column).value

    @property
    def row_count(self) -> int:
        """Number of rows (including header if given) the excel file has"""
        return self._shape["row_count"]

    @property
    def column_count(self) -> int:
        """Number of columns the excel file has"""
        return self._shape["column_count"]

    def load_all_data_from_workbook(self, tab_name: str = None) -> list[dict]:
        """Loads the data from a tab in a workbook"""
        wb_obj = openpyxl.load_workbook(
            filename=io.BytesIO(self._contents), data_only=True
        )
        # self._sheet_obj = wb_obj[tab_name] if tab_name else wb_obj["sheet1"]       if tab_name:
        self._sheet_obj = wb_obj[tab_name] if tab_name else wb_obj.active
        self._shape["row_count"] = self._sheet_obj.max_row
        self._shape["column_count"] = self._sheet_obj.max_column
        return self._sheet_obj

    def to_data_frame(self) -> pd.DataFrame:
        column_names = []
        data_rows = []
        for indx, row in enumerate(self._sheet_obj):
            if indx == 0:
                # Header row
                for cell in row:
                    column_names.append(cell.value)
            else:
                # Data row
                data_row = []
                for cell in row:
                    data_row.append(cell.value)
                data_rows.append(data_row)

        return pd.DataFrame(data_rows, columns=column_names)


class CSV_Raw_Stream(StreamUtil):
    """Local loading and local writing of Raw CSV files stream"""

    def load_data(self, csv_name: str, **options) -> list[str]:
        self._missing_values: list[str] = options.get("missing_values", [])
        with open(csv_name, mode="r") as file:
            self._df_data = list(csv.reader(file))
            self._header = self._df_data[0]
            self._df_data = self._df_data[1:]
            return self._df_data
        return None

    @property
    def row_count(self) -> int:
        """Number of rows (including header if given) the csv file has"""
        return len(self._df_data)

    @property
    def column_count(self) -> int:
        """Number of columns the csv file has"""
        return len(self._df_data[0])

    def get_column_names(self) -> list[str]:
        """Holds a list of all the column names"""
        return self._header

    def get_column_name(self, column: int) -> str:
        """Get the column name from column_no (1 based)"""
        return self.get_column_names()[column - 1]

    def get_row_column_contents(self, row: int, column: int):
        """Get the value from row (1 based), column (1 based)"""
        return str(self._df_data[row - 1][column - 1])

    @staticmethod
    def _preprocess_application_data(data: Any, missing_values: list[str]) -> Any:
        """Changes the data to empty when any missing values are found"""
        if data in missing_values:
            return None
        return data


class CSV_DataFrame_Stream(StreamUtil):
    """Local or remote loading and local writing of CSV files stream
    that uses Pandas to match the CSV data to use as a DataFrame"""

    def load_data(self, csv_name: str, **options) -> pd.DataFrame:
        self._missing_values: list[str] = options.get("missing_values", [])
        self._df_data = pd.read_csv(csv_name, na_values=self._missing_values)
        return self._df_data

    def save_data_to_csv(self, csv_name: str) -> None:
        """Save the loaded csv file to local file (used for remote data loads)"""
        self._df_data.to_csv(csv_name, index=None)

    @property
    def row_count(self) -> int:
        """Number of rows (including header if given) the csv file has"""
        return self._df_data.index.stop

    @property
    def column_count(self) -> int:
        """Number of columns the csv file has"""
        return len(self._df_data.columns)

    def get_column_names(self) -> list[str]:
        """Holds a all the column name"""
        return self._df_data.columns

    def get_column_name(self, column: int) -> str:
        """Get the column name from column_no (1 based)"""
        return self.get_column_names()[column - 1]

    def get_row_column_contents(self, row: int, column: int):
        """Get the value from row (1 based), column (1 based)"""
        return str(self._df_data.iat[row - 1, column - 1])

    @staticmethod
    def _preprocess_application_data(data: Any, missing_values: list[str]) -> Any:
        """Changes the data to empty when any missing values are found"""
        if data in missing_values:
            return None
        return data

    def pandas_style_to_data_frame(self):
        """WIP: When the worksheet does have headers or indices, such as one created by Pandas"""
        data = self._sheet_obj.values
        cols = next(data)[1:]
        data = list(data)
        idx = [r[0] for r in data]
        data = (it.islice(r, 1, None) for r in data)
        return pd.DataFrame(data, index=idx, columns=cols)


class CSVFactory(DataDesc):
    """CSV Factory to read CSV files and transform to a specified data format"""

    def __init__(self, **meta) -> None:
        super().__init__(**meta)
        self._Stream = None

    def create_dataframe(self, csv_name: str, **options) -> pd.DataFrame:
        """Creates a memory based DataFrame data format (for use in Pandas) from the loaded csv file"""
        self._Stream = CSV_DataFrame_Stream()
        self._data = self._Stream.load_data(csv_name, **options)
        return self._data

    def create_column_key(self, csv_name: str, **options) -> list[ColumnKey]:
        """Creates a memory based list of Column data identified by Key data format from the loaded csv file"""
        self._Stream = CSV_Raw_Stream()
        self._Stream.load_data(csv_name, **options)
        self._data = self._Stream.to_column_key()
        return self._data

    @property
    def row_count(self):
        """Number of rows the read table has"""
        return self._Stream.row_count

    @property
    def column_count(self):
        """Number of columns the read table has"""
        return self._Stream.column_count


class ExcelFactory(DataDesc):
    """Excel Factory to read Excel xlsx files and transform to a specificied data format"""

    def __init__(self, **meta) -> None:
        super().__init__(**meta)
        self._Stream = None

    def create_dataframe(
        self, excel_name: str, tab_name: str = None, **options
    ) -> pd.DataFrame:
        """Creates a memory based DataFrame data format (for use in Pandas) from the loaded xlsx file"""
        self._Stream = Local_Excel_Workbook_Stream()
        self._Stream.load_data(excel_name, **options)
        self._Stream.load_all_data_from_workbook(tab_name)
        self._data = self._Stream.to_data_frame()
        return self._data

    def create_column_key(
        self, excel_name: str, tab_name: str = None, **options
    ) -> list[ColumnKey]:
        """Creates a memory based list of Column data identified by Key data format from the loaded xlsx file"""
        self._Stream = Local_Excel_Workbook_Stream()
        self._Stream.load_data(excel_name, **options)
        self._Stream.load_all_data_from_workbook(tab_name)
        self._data = self._Stream.to_column_key()
        return self._data

    @property
    def row_count(self):
        """Number of rows the read table has"""
        return self._Stream.row_count

    @property
    def column_count(self):
        """Number of columns the read table has"""
        return self._Stream.column_count


if __name__ == "__main__":
    factory = ExcelFactory()
    data = factory.create_column_key("sales_data_types.xlsx", missing_values=["[NULL]"])
    print(data)
    print(factory.row_count, factory.column_count)

    df = factory.create_dataframe(
        "sales_data_types.xlsx", "sales_data_types", missing_values=["[NULL]"]
    )
    print(factory.row_count, factory.column_count)
    # Convert the entire DataFrame
    print(df.to_csv())
    print(df.to_numpy())

    factory = CSVFactory()
    data = factory.create_column_key("sales_data_types.csv", missing_values=["[NULL]"])
    print(factory.row_count, factory.column_count)

    df = factory.create_dataframe("sales_data_types.csv", missing_values=["[NULL]"])
    print(factory.row_count, factory.column_count)
    print(df.to_csv())

    # Convert the entire DataFrame
    print(df.to_numpy())

    # Convert specific columns
    # df_fact.data[['A', 'C']].to_numpy()

    # Get Numpy array with dtypes in the result
    print(df.to_records())
